import os, re, json, hashlib
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

ROOT              = os.path.dirname(__file__)
card_base         = os.path.join(ROOT, "..", "Code", "Cards")
power_base        = os.path.join(ROOT, "..", "Code", "Powers")
power_img_base    = os.path.join(ROOT, "powers")
card_loc          = os.path.join(ROOT, "..", "Downfall", "localization", "eng", "cards.json")
power_loc         = os.path.join(ROOT, "..", "Downfall", "localization", "eng", "powers.json")
PLACEHOLDERS_FILE = os.path.join(ROOT, ".placeholders.json")
THUMB_CACHE_DIR   = os.path.join(ROOT, ".excel_cache")
OUT_EXCEL         = os.path.join(ROOT, "missing_assets.xlsx")

CARD_DIRS = [
    os.path.join(ROOT, "cards"),
    os.path.join(ROOT, "cards_beta"),
    os.path.join(ROOT, "cards_missing"),
]

THUMB_H       = 57
ROW_HEIGHT_PT = THUMB_H * 0.75

os.makedirs(THUMB_CACHE_DIR, exist_ok=True)

# ── Helpers ───────────────────────────────────────────────────

def to_snake(name):
    return re.sub(r'(?<!^)(?=[A-Z])', '_', name).lower()

def normalize(name):
    return to_snake(name).replace("_", "")

def file_hash(path):
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()

def load_loc(path, suffix=""):
    data = {}
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)
    for key, value in raw.items():
        key   = key.removeprefix("DOWNFALL-")
        parts = key.split(".", 1)
        name  = parts[0]
        field = parts[1] if len(parts) > 1 else ""
        if suffix:
            name = re.sub(rf'_{re.escape(suffix)}$', '', name, flags=re.IGNORECASE)
        norm = normalize(name)
        if norm not in data:
            data[norm] = {}
        data[norm][field] = value
    return data

# ── Placeholders ──────────────────────────────────────────────

placeholders = {}
if os.path.exists(PLACEHOLDERS_FILE):
    with open(PLACEHOLDERS_FILE) as f:
        placeholders = json.load(f)

def is_placeholder(path):
    rel = os.path.relpath(path, ROOT)
    return rel in placeholders and file_hash(path) == placeholders[rel]

# ── Thumbnails ────────────────────────────────────────────────

def get_thumbnail(img_path):
    src_hash   = file_hash(img_path)
    thumb_path = os.path.join(THUMB_CACHE_DIR, src_hash + ".png")
    if os.path.exists(thumb_path):
        return thumb_path

    img = Image.open(img_path).convert("RGBA")

    # Composite onto white so alpha shows correctly in Excel
    bg = Image.new("RGBA", img.size, (255, 255, 255, 255))
    bg.paste(img, mask=img.split()[3])
    img = bg.convert("RGB")

    ratio = img.width / img.height
    w = max(1, int(THUMB_H * ratio))
    img = img.resize((w, THUMB_H), Image.LANCZOS)
    img.save(thumb_path, format="PNG")
    return thumb_path

# ── Image lookup ──────────────────────────────────────────────

def card_img_lookup(rel_folder, stem):
    subfolder = "" if rel_folder in (".", "") else rel_folder
    for base in CARD_DIRS:
        if not os.path.exists(base):
            continue
        # Try exact path first
        candidate = os.path.join(base, subfolder, f"{stem}.png")
        if os.path.exists(candidate):
            return candidate
        # Case-insensitive fallback — walk all subdirs
        for root, dirs, files in os.walk(base):
            for file in files:
                if file.lower() == f"{stem}.png":
                    return os.path.join(root, file)
    return None

def power_img_lookup(rel_folder, stem):
    for root, dirs, files in os.walk(power_img_base):
        for file in files:
            if file.lower() == f"{stem}.png":
                return os.path.join(root, file)
    return None

# ── Collect cards ─────────────────────────────────────────────

cards, card_images = defaultdict(dict), defaultdict(set)

for root, dirs, files in os.walk(card_base):
    dirs[:] = [d for d in dirs if d != "Abstract"]
    rel    = os.path.relpath(root, card_base)
    folder = normalize(rel.split(os.sep)[0])
    for file in files:
        if file.endswith(".cs"):
            snake = to_snake(os.path.splitext(file)[0])
            cards[folder][normalize(snake)] = (snake, rel)

for base in CARD_DIRS:
    if not os.path.exists(base):
        continue
    for root, dirs, files in os.walk(base):
        folder = normalize(os.path.relpath(root, base).split(os.sep)[0])
        for file in files:
            if file.endswith(".png"):
                path = os.path.join(root, file)
                if not is_placeholder(path):
                    card_images[folder].add(normalize(os.path.splitext(file)[0]))

# ── Collect powers ────────────────────────────────────────────

powers, power_images = defaultdict(dict), defaultdict(set)

for root, dirs, files in os.walk(power_base):
    dirs[:] = [d for d in dirs if d != "Abstract"]
    folder = normalize(os.path.relpath(root, power_base).split(os.sep)[0])
    for file in files:
        if file.endswith(".cs"):
            snake = to_snake(os.path.splitext(file)[0])
            snake = re.sub(r'_power$', '', snake)
            powers[folder][normalize(snake)] = (snake, ".")

for root, dirs, files in os.walk(power_img_base):
    folder = normalize(os.path.relpath(root, power_img_base).split(os.sep)[0])
    for file in files:
        if file.endswith(".png"):
            path = os.path.join(root, file)
            if not is_placeholder(path):
                power_images[folder].add(normalize(os.path.splitext(file)[0]))

card_locs  = load_loc(card_loc)
power_locs = load_loc(power_loc, suffix="power")

# ── Excel ─────────────────────────────────────────────────────

wb     = Workbook()
RED    = PatternFill("solid", fgColor="FFCCCC")
GREEN  = PatternFill("solid", fgColor="CCFFCC")
HEADER = PatternFill("solid", fgColor="4472C4")
BOLD   = Font(bold=True, color="FFFFFF")

def make_sheet(wb, title, code_dict, image_dict, loc, img_lookup_fn):
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 12

    for col, text in enumerate(["Folder", "File Name", "Title", "Description", "Image"], 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = HEADER
        cell.font = BOLD

    def write_row(row, folder, norm, fill):
        snake, rel_folder = code_dict[folder][norm]
        loc_entry = loc.get(norm, {})
        title_str = loc_entry.get("title", "")
        desc_str  = re.sub(r'\[/?[^\]]+\]', '', loc_entry.get("description", "")).strip()

        ws.cell(row=row, column=1, value=folder).fill    = fill
        ws.cell(row=row, column=2, value=snake).fill     = fill
        ws.cell(row=row, column=3, value=title_str).fill = fill
        ws.cell(row=row, column=4, value=desc_str).fill  = fill
        ws.row_dimensions[row].height = ROW_HEIGHT_PT

        img_path = img_lookup_fn(rel_folder, snake)
        if img_path:
            try:
                thumb  = get_thumbnail(img_path)
                pil    = Image.open(thumb)
                xl_img = XLImage(thumb)
                xl_img.width  = pil.width
                xl_img.height = pil.height
                ws.add_image(xl_img, f"E{row}")
            except Exception as e:
                ws.cell(row=row, column=5, value=f"err: {e}")

    row = 2
    all_folders = sorted(set(list(code_dict.keys()) + list(image_dict.keys())))

    # Missing (red)
    for folder in all_folders:
        missing = set(code_dict.get(folder, {}).keys()) - image_dict.get(folder, set())
        for norm in sorted(missing):
            write_row(row, folder, norm, RED)
            row += 1

    # Has art (green)
    for folder in all_folders:
        has_art = set(code_dict.get(folder, {}).keys()) & image_dict.get(folder, set())
        for norm in sorted(has_art):
            write_row(row, folder, norm, GREEN)
            row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="All good ✓")

del wb["Sheet"]
make_sheet(wb, "Cards",  cards,  card_images, card_locs,  card_img_lookup)
make_sheet(wb, "Powers", powers, power_images, power_locs, power_img_lookup)

wb.save(OUT_EXCEL)
print(f"Saved: {OUT_EXCEL}")